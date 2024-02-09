from openpyxl import Workbook, load_workbook


wb = load_workbook("Titanic_models.xlsx")

ws = wb.active

sheetCreate = wb.create_sheet("demo1")

print(sheetCreate)