from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# create a sheet
sheet1 = wb.create_sheet("Sheet2")
ws.title = "Sheet1"

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)

print(wb.sheetnames)

print(target)
